"""Reference tests for `device_data_exporter` scaffolding."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest

from references.device_data_util import device_data_exporter as exporter


class TestSelectExportColumns:
    def test_returns_union_of_keys(self) -> None:
        """Column selection should union keys when include/exclude are absent."""

        rows = [{"name": "core", "serial": "ABC"}, {"platform": "ios", "name": "edge"}]

        columns = exporter.select_export_columns(rows)

        assert columns == ["name", "platform", "serial"]

    def test_honors_include_whitelist(self) -> None:
        """Explicit include sequences should control the final column order."""

        rows = [{"name": "core", "serial": "ABC", "platform": "ios"}]

        columns = exporter.select_export_columns(rows, include=["serial", "name"])

        assert columns == ["serial", "name"]

    def test_exclude_columns_are_removed(self) -> None:
        """Excluded columns should be filtered from the result."""

        rows = [{"name": "core", "serial": "ABC", "platform": "ios"}]

        columns = exporter.select_export_columns(rows, exclude=["serial"])

        assert columns == ["name", "platform"]


class TestRenderCsv:
    def test_serializes_rows_to_csv_string(self) -> None:
        """Rows should render into a CSV-formatted text blob."""

        rows = [{"name": "core", "serial": "ABC"}]
        columns = ["name", "serial"]

        csv_text = exporter.render_csv(rows, columns=columns)

        assert csv_text == "name,serial\ncore,ABC\n"

    def test_custom_newline_is_supported(self) -> None:
        """Callers can override newline handling for platform compatibility."""

        rows = [{"name": "core", "serial": "ABC"}]
        columns = ["name", "serial"]

        csv_text = exporter.render_csv(rows, columns=columns, newline="\r\n")

        assert csv_text == "name,serial\r\ncore,ABC\r\n"


class TestStreamCsv:
    def test_yields_header_and_rows(self) -> None:
        """Streaming CSV should yield header first followed by each row."""

        rows = [{"name": "core", "serial": "ABC"}, {"name": "edge", "serial": "XYZ"}]
        columns = ["name", "serial"]

        chunks = list(exporter.stream_csv(rows, columns=columns))

        assert chunks == ["name,serial\n", "core,ABC\n", "edge,XYZ\n"]


class TestRenderJson:
    def test_outputs_compact_json_by_default(self) -> None:
        """JSON rendering should default to compact formatting."""

        rows = [{"name": "core"}]

        payload = exporter.render_json(rows)

        assert payload == '[{"name": "core"}]'

    def test_supports_indentation(self) -> None:
        """Callers can request pretty-printed JSON with an indent value."""

        rows = [{"name": "core"}]

        payload = exporter.render_json(rows, indent=2)

        assert payload == '[\n  {\n    "name": "core"\n  }\n]'


class TestWriteExcel:
    def test_creates_excel_file_with_expected_sheet(self, tmp_path: Path) -> None:
        """Excel writer should persist records into the requested worksheet."""

        destination = tmp_path / "devices.xlsx"
        rows = [{"name": "core", "serial": "ABC"}]
        columns = ["name", "serial"]

        exporter.write_excel(rows, columns=columns, destination=destination, worksheet_title="Devices")

        from openpyxl import load_workbook

        wb = load_workbook(destination)
        ws = wb["Devices"]
        assert ws.max_row == 2
        assert ws.max_column == 2
        assert [cell.value for cell in ws[1]] == ["name", "serial"]
        assert [cell.value for cell in ws[2]] == ["core", "ABC"]


class TestMakeExportFilename:
    def test_uses_timestamp_when_provided(self) -> None:
        """Generated filenames should include formatted timestamps."""

        timestamp = datetime(2023, 1, 2, 3, 4, 5)

        name = exporter.make_export_filename("devices", extension="csv", timestamp=timestamp)

        assert name == "devices_20230102030405.csv"

    def test_defaults_to_current_time(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Timestamp should default to `datetime.utcnow` when absent."""

        fake_now = datetime(2023, 7, 8, 9, 10, 11)

        class _FakeDateTime(datetime):
            @classmethod
            def utcnow(cls) -> datetime:
                return fake_now

        monkeypatch.setattr(exporter, "datetime", _FakeDateTime)

        name = exporter.make_export_filename("devices", extension="json")

        assert name == "devices_20230708091011.json"


class TestSummarizeExport:
    def test_reports_record_count_and_columns(self) -> None:
        """Export summary should expose record and column metadata."""

        rows = [{"name": "core"}, {"name": "edge"}]
        columns = ["name"]

        summary = exporter.summarize_export(rows, columns=columns)

        assert summary["total_rows"] == 2
        assert summary["columns"] == columns


def test_write_excel_accepts_filelike_destination(tmp_path: Path) -> None:
    """Utility should support writing to open binary handles as well as paths."""

    buffer = tmp_path / "devices.xlsx"
    with buffer.open("wb") as handle:
        exporter.write_excel([
            {"name": "core", "serial": "ABC"},
            {"name": "edge", "serial": "XYZ"},
        ], columns=["name", "serial"], destination=handle)

    from openpyxl import load_workbook

    wb = load_workbook(buffer)
    ws = wb.active
    assert [cell.value for cell in ws[1]] == ["name", "serial"]
